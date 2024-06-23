from openpyxl import load_workbook
import os
import platform
import subprocess
import pandas as pd

# def ping_ip(ip):
#     # Option for the number of packers as a function of OS
#     param = '-n' if platform.system().lower() == 'windows' else '-c'
#     command = ['ping', param, '1', ip]
#     # response = os.system(f"ping -c 3 {ip}")
#     return os.system(' '.join(command)) == 0

def ping_ip_subprocess_Exists(ip):
    # Option for the number of packers as a function of OS
    param = '-n' if platform.system().lower() == 'windows' else '-c'
    command = ['ping', param, '1', ip]
    response = subprocess.run(command, capture_output=True)
    return response.returncode == 0

def modifyExcelByPing(file_path, read_column, write_column):
    workbook = load_workbook(filename=file_path)

    sheet = workbook.active
    for row in range(2, sheet.max_row):
        ip = sheet.cell(row=row, column=read_column).value
        if ip is not None:
            status = 'online' if ping_ip_subprocess_Exists(ip) else 'offline'
            print(f'{row} --- {ip} : {status}')
            sheet.cell(row=row, column=write_column, value=status)
        else:
            status = 'Not valid IP'
            print(status)
            sheet.cell(row=row, column=write_column, value=status)
    workbook.save(f'{file_path.split(".")[0]}_mod.{file_path.split(".")[1]}')

def pingXl2csv(file_path, read_column, write_column):
    df = pd.read_excel(file_path)
    # print(df.head())
    for index in range(1, len(df)):
        ip = str(df.iloc[index, read_column])
        # print(type(ip))
        if ip != 'nan':
            status = 'online' if ping_ip_subprocess_Exists(ip) else 'offline'
            print(f'{index} --- {ip} : {status}')   
        else:
            status = 'Not valid IP'
            print(status)

        df.iloc[index, write_column] = status
    df.to_csv(f'{file_path.split(".")[0]}_mod.csv')

if __name__ == "__main__":
    pingXl2csv("all_veam_backup_list_6-22.xlsm", 2, 3)